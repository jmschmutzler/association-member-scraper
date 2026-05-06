import io
import pytest
from openpyxl import load_workbook
from extractor import Company
from writer import company_to_row, deduplicate, write_csv, write_xlsx, FIELDS


def _make_company(name="Acme Corp", website="https://acme.com"):
    return Company(
        name=name,
        website=website,
        email=None,
        phone=None,
        address=None,
        industry=None,
        description=None,
    )


def test_company_to_row_sets_source_url():
    row = company_to_row(_make_company(), source_url="https://dir.com/members")
    assert row.source_url == "https://dir.com/members"


def test_company_to_row_sets_scraped_at():
    row = company_to_row(_make_company(), source_url="https://dir.com")
    assert row.scraped_at  # non-empty ISO string


def test_deduplicate_removes_duplicate_name_and_website():
    r1 = company_to_row(_make_company("Acme Corp", "https://acme.com"), "src1")
    r2 = company_to_row(_make_company("Acme Corp", "https://acme.com"), "src2")
    result = deduplicate([r1, r2])
    assert len(result) == 1


def test_deduplicate_last_write_wins():
    r1 = company_to_row(_make_company("Acme Corp", "https://acme.com"), "src1")
    r2 = company_to_row(_make_company("Acme Corp", "https://acme.com"), "src2")
    result = deduplicate([r1, r2])
    assert result[0].source_url == "src2"


def test_deduplicate_keeps_different_companies():
    r1 = company_to_row(_make_company("Acme Corp", "https://acme.com"), "src")
    r2 = company_to_row(_make_company("Beta LLC", "https://beta.com"), "src")
    result = deduplicate([r1, r2])
    assert len(result) == 2


def test_deduplicate_is_case_insensitive_on_name():
    r1 = company_to_row(_make_company("acme corp", "https://acme.com"), "src1")
    r2 = company_to_row(_make_company("Acme Corp", "https://acme.com"), "src2")
    result = deduplicate([r1, r2])
    assert len(result) == 1


def test_write_csv_includes_header():
    rows = [company_to_row(_make_company(), "https://dir.com")]
    csv_str = write_csv(rows)
    first_line = csv_str.splitlines()[0]
    assert first_line == ",".join(FIELDS)


def test_write_csv_includes_company_data():
    rows = [company_to_row(_make_company("Acme Corp", "https://acme.com"), "https://dir.com")]
    csv_str = write_csv(rows)
    assert "Acme Corp" in csv_str
    assert "https://acme.com" in csv_str


def test_write_xlsx_has_companies_sheet_with_data():
    rows = [company_to_row(_make_company("Acme Corp", "https://acme.com"), "https://dir.com")]
    result = write_xlsx(rows, [])
    assert isinstance(result, bytes)
    wb = load_workbook(io.BytesIO(result))
    assert "Companies" in wb.sheetnames
    ws = wb["Companies"]
    header = [cell.value for cell in ws[1]]
    assert header == FIELDS
    assert ws.cell(row=2, column=1).value == "Acme Corp"


def test_write_xlsx_has_errors_sheet():
    errors = [{"url": "https://fail.com", "reason": "timeout"}]
    result = write_xlsx([], errors)
    wb = load_workbook(io.BytesIO(result))
    assert "_errors" in wb.sheetnames


def test_write_xlsx_errors_sheet_contains_failed_url():
    errors = [{"url": "https://fail.com", "reason": "timeout"}]
    result = write_xlsx([], errors)
    wb = load_workbook(io.BytesIO(result))
    ws = wb["_errors"]
    urls = [row[0].value for row in ws.iter_rows(min_row=2)]
    assert "https://fail.com" in urls
